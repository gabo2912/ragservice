"""
orquestador.py — Modo "Conversa conmigo" gobernado por Gemini (agente con tools).

A diferencia del pipeline de reglas del chatbot, aquí Gemini ORQUESTA la
conversación libre. En cada turno decide, mediante function calling, qué
herramienta usar:

    1. traducir_palabra(palabra)        → corpus de vocabulario (palabras.xlsx)
    2. buscar_frase_conversacional(...)  → frases_conversacionales.xlsx
    3. consultar_cultura(pregunta)       → RAG sobre el PDF de cosmovisión
    4. info_aplicacion(tema)             → qué es Pishico y cómo se usa

Si el mensaje es charla general (saludos, "¿cómo estás?", agradecer), Gemini
responde directamente sin herramienta.

REGLA ANTI-ALUCINACIÓN (crítica para un proyecto de revitalización):
Gemini NUNCA inventa palabras ni frases en shipibo-konibo. Para CUALQUIER dato
en shipibo debe llamar a una herramienta; si la herramienta no lo encuentra,
lo dice honestamente. El modelo conversa y decide, pero los datos lingüísticos
salen siempre de los archivos curados, nunca de su imaginación.

Si Gemini no está disponible (sin API key, sin SDK, rate limit agotado), el
endpoint devuelve una señal para que el chatbot use su pipeline de reglas
clásico como respaldo. Nunca se rompe la conversación.
"""

import logging
import os
import time
import unicodedata
from pathlib import Path
from typing import Any, Dict, List, Optional

from . import config, responder

logger = logging.getLogger(__name__)

# ── Rutas de datos ───────────────────────────────────────────────────────────
# Los archivos de corpus y frases se cargan localmente en el rag-service.
# Se pueden sobreescribir por variable de entorno; por defecto se buscan en
# data/ junto al índice.
_DATA_DIR = Path(os.getenv("ORQ_DATA_DIR", Path(__file__).resolve().parent.parent / "data"))
_PALABRAS_XLSX = Path(os.getenv("ORQ_PALABRAS_XLSX", _DATA_DIR / "palabras.xlsx"))
_FRASES_XLSX = Path(os.getenv("ORQ_FRASES_XLSX", _DATA_DIR / "frases_conversacionales.xlsx"))


# ── Normalización ────────────────────────────────────────────────────────────
def _norm(s: str) -> str:
    if not s:
        return ""
    t = str(s).lower().strip()
    t = "".join(ch for ch in unicodedata.normalize("NFD", t)
                if unicodedata.category(ch) != "Mn")
    return t


# ── Carga perezosa de datos (una sola vez) ───────────────────────────────────
_palabras: Optional[List[Dict[str, str]]] = None
_frases: Optional[List[Dict[str, str]]] = None


def _cargar_palabras() -> List[Dict[str, str]]:
    global _palabras
    if _palabras is not None:
        return _palabras
    _palabras = []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(_PALABRAS_XLSX, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 3:
                continue
            es, shp = row[1], row[2]
            if es and shp:
                _palabras.append({"es": str(es).strip(), "shp": str(shp).strip()})
        wb.close()
        logger.info("orquestador: %d palabras cargadas desde %s", len(_palabras), _PALABRAS_XLSX)
    except FileNotFoundError:
        logger.warning("orquestador: no se encontró %s; traducción no disponible", _PALABRAS_XLSX)
    except Exception as e:
        logger.warning("orquestador: error cargando palabras: %s", e)
    return _palabras


def _cargar_frases() -> List[Dict[str, str]]:
    global _frases
    if _frases is not None:
        return _frases
    _frases = []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(_FRASES_XLSX, data_only=True, read_only=True)
        ws = wb["frases"] if "frases" in wb.sheetnames else wb[wb.sheetnames[0]]
        headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {h: i for i, h in enumerate(headers)}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            def get(col):
                i = idx.get(col)
                return str(row[i]).strip() if i is not None and i < len(row) and row[i] else ""
            es, shp, cat = get("es"), get("shp"), get("categoria")
            if es and shp:
                _frases.append({"es": es, "shp": shp, "categoria": cat, "tipo": get("tipo")})
        wb.close()
        logger.info("orquestador: %d frases cargadas desde %s", len(_frases), _FRASES_XLSX)
    except FileNotFoundError:
        logger.warning("orquestador: no se encontró %s; frases no disponibles", _FRASES_XLSX)
    except Exception as e:
        logger.warning("orquestador: error cargando frases: %s", e)
    return _frases


# ═════════════════════════════════════════════════════════════════════════════
# HERRAMIENTAS (las ejecuta el orquestador cuando Gemini las invoca)
# ═════════════════════════════════════════════════════════════════════════════

def tool_traducir_palabra(palabra: str) -> Dict[str, Any]:
    """Traduce una palabra puntual usando el corpus curado. No inventa."""
    palabras = _cargar_palabras()
    n = _norm(palabra)
    # match exacto es→shp o shp→es
    for p in palabras:
        if _norm(p["es"]) == n:
            return {"encontrada": True, "es": p["es"], "shp": p["shp"], "direccion": "es_a_shp"}
    for p in palabras:
        # el shp puede traer nota entre paréntesis: "yapa (pescado vivo)"
        shp_base = p["shp"].split("(")[0].strip()
        if _norm(shp_base) == n:
            return {"encontrada": True, "es": p["es"], "shp": shp_base, "direccion": "shp_a_es"}
    return {"encontrada": False, "palabra": palabra}


def tool_buscar_frase_conversacional(situacion: str) -> Dict[str, Any]:
    """Busca frases útiles para una situación (saludo, despedida, agradecer,
    cortesia, ayuda, afirmacion, negacion, identidad, emocion)."""
    frases = _cargar_frases()
    n = _norm(situacion)
    # mapa de sinónimos → categoría del archivo
    alias = {
        "saludar": "saludo", "saludo": "saludo", "hola": "saludo", "buenos dias": "saludo",
        "despedir": "despedida", "despedida": "despedida", "chau": "despedida", "adios": "despedida",
        "agradecer": "agradecer", "gracias": "agradecer",
        "cortesia": "cortesia", "por favor": "cortesia", "permiso": "cortesia",
        "ayuda": "ayuda", "ayudar": "ayuda", "socorro": "ayuda",
        "afirmar": "afirmacion", "afirmacion": "afirmacion", "si": "afirmacion",
        "negar": "negacion", "negacion": "negacion", "no": "negacion",
        "identidad": "identidad", "presentarme": "identidad", "nombre": "identidad",
        "emocion": "emocion", "sentimiento": "emocion", "feliz": "emocion", "triste": "emocion",
    }
    cat = alias.get(n)
    if not cat:
        # buscar coincidencia parcial en el texto de la situación
        for k, v in alias.items():
            if k in n:
                cat = v
                break
    seleccion = [f for f in frases if _norm(f["categoria"]) == _norm(cat or "")]
    if not seleccion:
        return {"encontradas": False, "situacion": situacion}
    return {
        "encontradas": True,
        "categoria": cat,
        "frases": [{"es": f["es"], "shp": f["shp"]} for f in seleccion[:6]],
    }


def tool_consultar_cultura(pregunta: str) -> Dict[str, Any]:
    """Consulta el conocimiento cultural (cosmovisión shipibo) vía RAG.
    Reutiliza responder.responder() en modo llm, que ya es anti-alucinación."""
    try:
        result = responder.responder(pregunta, modo="llm")
        resp = result.get("respuesta")
        if resp:
            return {"encontrada": True, "respuesta": resp}
        return {"encontrada": False, "pregunta": pregunta}
    except Exception as e:
        logger.warning("orquestador: consultar_cultura falló: %s", e)
        return {"encontrada": False, "pregunta": pregunta, "error": str(e)}


_INFO_APP = {
    "general": (
        "Pishico es un asistente educativo bilingüe shipibo-konibo y español. "
        "Ayuda a practicar vocabulario, leer cuentos tradicionales, conocer la "
        "cultura shipibo y conversar de forma libre."
    ),
    "modulos": (
        "Pishico tiene cuatro modos: Vocabulario (aprender y evaluar palabras por "
        "categorías con imágenes), Cuentos (relatos interactivos con preguntas), "
        "Conversar (este modo libre), y Mi Aprendizaje (tu progreso)."
    ),
    "vocabulario": (
        "En Vocabulario practicas palabras organizadas por categorías (animales, "
        "colores, cuerpo, naturaleza, objetos, números, familia). Primero puedes "
        "aprenderlas con imágenes y luego evaluarte."
    ),
    "cuentos": (
        "En Cuentos lees relatos tradicionales shipibo con preguntas intercaladas "
        "para practicar la comprensión."
    ),
    "objetivo": (
        "Pishico busca apoyar la práctica oral y la valoración de la lengua "
        "shipibo-konibo, contribuyendo a su revitalización."
    ),
}


def tool_info_aplicacion(tema: str) -> Dict[str, Any]:
    """Información sobre qué es Pishico y cómo usarlo."""
    n = _norm(tema)
    for clave, texto in _INFO_APP.items():
        if clave in n:
            return {"tema": clave, "info": texto}
    # por defecto, info general
    return {"tema": "general", "info": _INFO_APP["general"]}


def tool_sugerir_modulo(modulo: str) -> Dict[str, Any]:
    """Sugiere ir a un módulo específico (vocabulario o cuentos) cuando el
    usuario quiere APRENDER vocabulario o LEER cuentos. El modo conversar no
    enseña esos contenidos; solo ofrece la redirección al módulo correcto.

    Devuelve el módulo normalizado y un mensaje de invitación. El botón real
    de navegación lo agrega el chatbot (actions.py) a partir de esta señal."""
    n = _norm(modulo)
    if "vocab" in n or "palabra" in n or "aprender palabra" in n:
        return {
            "modulo": "vocabulario",
            "mensaje": "Para aprender vocabulario, te llevo al módulo de Vocabulario, "
                       "donde puedes practicar palabras por categorías con imágenes.",
        }
    if "cuento" in n or "relato" in n or "historia" in n or "leer" in n:
        return {
            "modulo": "cuentos",
            "mensaje": "Para leer cuentos tradicionales shipibo, te llevo al módulo de Cuentos.",
        }
    return {"modulo": None, "mensaje": ""}


# ═════════════════════════════════════════════════════════════════════════════
# ORQUESTADOR (Gemini con function calling)
# ═════════════════════════════════════════════════════════════════════════════

_PROMPT_ORQUESTADOR = (
    "Eres Pishico, un asistente conversacional cálido y respetuoso que ayuda a "
    "las personas a acercarse a la lengua y la cultura shipibo-konibo. Hablas en "
    "español, de forma natural, breve y amable.\n\n"
    "REGLA ABSOLUTA E INVIOLABLE: NUNCA inventes palabras, frases ni traducciones "
    "en shipibo-konibo. Tu conocimiento del shipibo NO es confiable. Para CUALQUIER "
    "dato en shipibo debes OBLIGATORIAMENTE usar una herramienta:\n"
    "- Para traducir una palabra puntual → usa traducir_palabra.\n"
    "- Para enseñar cómo decir algo en una situación (saludar, despedirse, "
    "agradecer, pedir ayuda, etc.) → usa buscar_frase_conversacional.\n"
    "- Para preguntas sobre cultura, cosmovisión, seres, plantas, historia "
    "shipibo → usa consultar_cultura.\n"
    "- Para preguntas sobre qué es esta aplicación o cómo usarla → usa "
    "info_aplicacion.\n\n"
    "- Si el usuario quiere APRENDER vocabulario, practicar palabras o LEER "
    "cuentos, NO se lo enseñes tú (este modo es solo conversación). En su lugar "
    "usa sugerir_modulo para ofrecerle ir al módulo correcto.\n\n"
    "Si una herramienta no encuentra el dato, dilo con honestidad ('no tengo esa "
    "palabra en mi vocabulario todavía') y NO lo inventes. Si el mensaje es charla "
    "general (saludo, cómo estás, gracias), responde tú directamente sin herramienta, "
    "con calidez. Mantén las respuestas cortas (2-4 oraciones). Al final de una "
    "respuesta útil, puedes invitar suavemente a seguir explorando."
)


def _tool_declarations():
    """Declaraciones de las herramientas para Gemini (formato google-genai)."""
    from google.genai import types
    return [
        types.Tool(function_declarations=[
            types.FunctionDeclaration(
                name="traducir_palabra",
                description="Traduce una palabra puntual entre español y shipibo-konibo usando el corpus curado. Úsala siempre que el usuario pida traducir o pregunte cómo se dice una palabra concreta.",
                parameters=types.Schema(
                    type=types.Type.OBJECT,
                    properties={"palabra": types.Schema(type=types.Type.STRING, description="La palabra a traducir, en español o shipibo")},
                    required=["palabra"],
                ),
            ),
            types.FunctionDeclaration(
                name="buscar_frase_conversacional",
                description="Devuelve frases útiles en shipibo para una situación conversacional: saludo, despedida, agradecer, cortesia, ayuda, afirmacion, negacion, identidad, emocion.",
                parameters=types.Schema(
                    type=types.Type.OBJECT,
                    properties={"situacion": types.Schema(type=types.Type.STRING, description="La situación, p.ej. 'saludar', 'despedirse', 'agradecer'")},
                    required=["situacion"],
                ),
            ),
            types.FunctionDeclaration(
                name="consultar_cultura",
                description="Consulta el conocimiento cultural shipibo (cosmovisión, seres, plantas medicinales, historia, río Ucayali, ayahuasca, etc.) a partir del documento de cosmovisión.",
                parameters=types.Schema(
                    type=types.Type.OBJECT,
                    properties={"pregunta": types.Schema(type=types.Type.STRING, description="La pregunta cultural del usuario")},
                    required=["pregunta"],
                ),
            ),
            types.FunctionDeclaration(
                name="info_aplicacion",
                description="Explica qué es la aplicación Pishico y cómo usarla (sus módulos: vocabulario, cuentos, conversar, objetivo).",
                parameters=types.Schema(
                    type=types.Type.OBJECT,
                    properties={"tema": types.Schema(type=types.Type.STRING, description="Tema: general, modulos, vocabulario, cuentos, objetivo")},
                    required=["tema"],
                ),
            ),
            types.FunctionDeclaration(
                name="sugerir_modulo",
                description="Ofrece redirigir al usuario al módulo de Vocabulario o de Cuentos cuando quiere aprender palabras o leer cuentos. Úsala en vez de intentar enseñar ese contenido tú mismo.",
                parameters=types.Schema(
                    type=types.Type.OBJECT,
                    properties={"modulo": types.Schema(type=types.Type.STRING, description="El módulo deseado: 'vocabulario' o 'cuentos'")},
                    required=["modulo"],
                ),
            ),
        ])
    ]


_DISPATCH = {
    "traducir_palabra": tool_traducir_palabra,
    "buscar_frase_conversacional": tool_buscar_frase_conversacional,
    "consultar_cultura": tool_consultar_cultura,
    "info_aplicacion": tool_info_aplicacion,
    "sugerir_modulo": tool_sugerir_modulo,
}


def _es_rate_limit(exc: Exception) -> bool:
    t = str(exc).lower()
    return "429" in t or "resource_exhausted" in t or "quota" in t or "rate limit" in t


def conversar(mensaje: str, historial: Optional[List[Dict[str, str]]] = None) -> Dict[str, Any]:
    """
    Procesa un turno del modo conversar con Gemini orquestando herramientas.

    Args:
        mensaje: texto del usuario.
        historial: lista de turnos previos [{"rol": "user"|"model", "texto": "..."}],
                   para memoria multi-turno. Opcional.

    Returns:
        {"respuesta": str|None, "disponible": bool, "herramientas_usadas": [...]}
        Si disponible=False, el caller debe usar su pipeline de reglas de respaldo.
    """
    client = responder._get_gemini_client()
    if client is None:
        return {"respuesta": None, "disponible": False, "herramientas_usadas": []}

    from google.genai import types

    # Construir el historial de contenidos
    contents = []
    for turno in (historial or []):
        rol = "user" if turno.get("rol") == "user" else "model"
        contents.append(types.Content(role=rol, parts=[types.Part(text=turno.get("texto", ""))]))
    contents.append(types.Content(role="user", parts=[types.Part(text=mensaje)]))

    tools = _tool_declarations()
    cfg = types.GenerateContentConfig(
        system_instruction=_PROMPT_ORQUESTADOR,
        tools=tools,
        temperature=0.4,
        http_options=types.HttpOptions(timeout=int(config.GEMINI_TIMEOUT * 1000)),
    )

    usadas: List[str] = []
    modulo_sugerido: Optional[str] = None  # 'vocabulario' | 'cuentos' si aplica
    MAX_TURNOS_TOOL = 4  # evita loops infinitos de tool-calling

    for _turno in range(MAX_TURNOS_TOOL):
        resp = None
        for intento in range(config.GEMINI_MAX_REINTENTOS + 1):
            try:
                resp = client.models.generate_content(
                    model=config.GEMINI_MODEL, contents=contents, config=cfg
                )
                break
            except Exception as e:
                if _es_rate_limit(e) and intento < config.GEMINI_MAX_REINTENTOS:
                    time.sleep(config.GEMINI_BACKOFF_BASE * (2 ** intento))
                    continue
                logger.warning("orquestador: Gemini falló (%s); respaldo a reglas", e)
                return {"respuesta": None, "disponible": False, "herramientas_usadas": usadas}
        if resp is None:
            return {"respuesta": None, "disponible": False, "herramientas_usadas": usadas}

        # ¿Gemini pidió llamar a una o más herramientas?
        fcalls = []
        cand = (resp.candidates or [None])[0]
        if cand and cand.content and cand.content.parts:
            for part in cand.content.parts:
                if getattr(part, "function_call", None):
                    fcalls.append(part.function_call)

        if not fcalls:
            # No hay tool call → respuesta final de texto
            texto = (resp.text or "").strip()
            return {"respuesta": texto or None, "disponible": True,
                    "herramientas_usadas": usadas, "modulo_sugerido": modulo_sugerido}

        # Ejecutar las herramientas y devolver resultados a Gemini
        contents.append(cand.content)  # el turno del modelo con las function_call
        tool_response_parts = []
        for fc in fcalls:
            nombre = fc.name
            args = dict(fc.args) if fc.args else {}
            usadas.append(nombre)
            fn = _DISPATCH.get(nombre)
            try:
                resultado = fn(**args) if fn else {"error": "herramienta desconocida"}
            except Exception as e:
                logger.warning("orquestador: tool %s falló: %s", nombre, e)
                resultado = {"error": str(e)}
            # Capturar el módulo sugerido para que el chatbot muestre el botón
            if nombre == "sugerir_modulo" and isinstance(resultado, dict):
                modulo_sugerido = resultado.get("modulo") or modulo_sugerido
            tool_response_parts.append(
                types.Part.from_function_response(name=nombre, response=resultado)
            )
        contents.append(types.Content(role="user", parts=tool_response_parts))

    # Se agotaron los turnos de tool-calling sin respuesta final
    logger.warning("orquestador: máximo de turnos de tool alcanzado")
    return {"respuesta": None, "disponible": False, "herramientas_usadas": usadas}